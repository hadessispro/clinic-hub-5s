#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Nhập bộ sổ Excel của kế toán vào Finance Vault.

Mặc định chạy ở chế độ THỬ: đọc, kiểm tra, đối chiếu, in báo cáo, không ghi gì.
Chỉ khi thêm --emit-sql mới sinh file SQL để nạp.

    python scripts/finance-import.py "Gui Bao"
    python scripts/finance-import.py "Gui Bao" --emit-sql finance-seed.sql

Thang kiểm tra năm tầng, sai tầng nào thì dừng ở tầng đó:
    1 cấu trúc · 2 kiểu dữ liệu · 3 nghiệp vụ · 4 bất biến kế toán · 5 đối chiếu chéo
"""
import argparse
import hashlib
import os
import re
import sys
from collections import defaultdict
from datetime import date, datetime

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit('Thiếu openpyxl. Cài bằng: pip install openpyxl')

# Console Windows mặc định dùng cp1252, không in được tiếng Việt có dấu.
# Ép UTF-8 để script chạy được ở mọi nơi mà không phải đặt biến môi trường.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding='utf-8', errors='replace')
    except (AttributeError, ValueError):
        pass


# ── tiện ích ────────────────────────────────────────────────────────────────

def money(v):
    return f'{v:,.0f}'


def sha256(path):
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(1 << 20), b''):
            h.update(chunk)
    return h.hexdigest()


def sql_str(v):
    if v is None or v == '':
        return 'null'
    return "'" + str(v).replace("'", "''") + "'"


def sql_date(v):
    if v is None:
        return 'null'
    if isinstance(v, datetime):
        v = v.date()
    return f"'{v.isoformat()}'"


class Report:
    def __init__(self):
        self.lines = []
        self.errors = []

    def w(self, s=''):
        self.lines.append(s)
        print(s)

    def fail(self, tier, msg):
        self.errors.append((tier, msg))

    def dump(self, path):
        with open(path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(self.lines))


# ── đọc danh mục ────────────────────────────────────────────────────────────

NATURE = {'Dư Nợ': 'debit', 'Dư Có': 'credit', 'Lưỡng tính': 'both'}


def read_accounts(path, rep):
    """Danh_sach_he_thong_tai_khoan.xlsx · tiêu đề dòng 3"""
    wb = load_workbook(path, data_only=True, read_only=True)
    sh = wb[wb.sheetnames[0]]
    rows = []
    for i, r in enumerate(sh.iter_rows(values_only=True), start=1):
        if i <= 3 or r is None or r[1] is None:
            continue
        code = str(r[1]).strip()
        if not code or not code[0].isdigit():
            continue
        nat = NATURE.get(str(r[3] or '').strip())
        if nat is None:
            rep.fail(3, f'Tài khoản {code}: tính chất "{r[3]}" không nhận diện được')
            nat = 'both'
        rows.append({
            'code': code,
            'name': str(r[2] or '').strip(),
            'name_en': str(r[4] or '').strip() or None,
            'nature': nat,
            'note': str(r[5] or '').strip() or None,
            'is_active': str(r[6] or '').strip() == 'Đang sử dụng',
        })
    wb.close()
    # cha con suy từ độ dài mã
    byc = {a['code'] for a in rows}
    for a in rows:
        a['depth'] = len(a['code']) - 2
        parent = a['code'][:-1]
        while parent and len(parent) >= 3 and parent not in byc:
            parent = parent[:-1]
        a['parent'] = parent if (parent != a['code'] and parent in byc) else None
    return rows


def read_cost_items(path, rep):
    """Danh_sach_khoan_muc_chi_phi.xlsx · mã dạng DN, DN.LVT, DN.PVC"""
    wb = load_workbook(path, data_only=True, read_only=True)
    sh = wb[wb.sheetnames[0]]
    rows = []
    for i, r in enumerate(sh.iter_rows(values_only=True), start=1):
        if i <= 3 or r is None or r[1] is None:
            continue
        code = str(r[1]).strip()
        if not code:
            continue
        branch = code.split('.')[1] if '.' in code else None
        rows.append({
            'code': code,
            'name': str(r[2] or '').strip(),
            'branch': branch,
            'is_active': str(r[4] or '').strip() == 'Đang sử dụng',
        })
    wb.close()
    return rows


def read_suppliers(path, rep):
    """Danh_sach_nha_cung_cap.xlsx · tiêu đề dòng 3"""
    wb = load_workbook(path, data_only=True, read_only=True)
    sh = wb[wb.sheetnames[0]]
    rows = []
    for i, r in enumerate(sh.iter_rows(values_only=True), start=1):
        if i <= 3 or r is None or r[1] is None:
            continue
        code = str(r[1]).strip()
        if not code:
            continue
        rows.append({
            'code': code,
            'name': str(r[2] or '').strip(),
            'kind': 'supplier',
            'address': str(r[3] or '').strip() or None,
            'tax_code': str(r[5] or '').strip() or None,
            'phone': str(r[8] or '').strip() or None,
        })
    wb.close()
    return rows


# tiền tố mã đối tượng cho biết loại
PARTNER_KIND = {'NCC': 'supplier', 'NV': 'employee'}
BRANCH_PREFIX = {'PVC', 'LVT'}


def classify_partner(code):
    pref = ''.join(ch for ch in code[:3] if ch.isalpha()).upper()
    kind = PARTNER_KIND.get(pref, 'customer')
    branch = pref if pref in BRANCH_PREFIX else None
    return kind, branch


# ── đọc nhật ký chung ───────────────────────────────────────────────────────

VOUCHER_TYPE = re.compile(r'^([A-Za-zĐđ]+)')


def read_journal(path, rep):
    wb = load_workbook(path, data_only=True, read_only=True)
    rows = []
    sheets = list(wb.sheetnames)
    for sname in sheets:
        for i, r in enumerate(wb[sname].iter_rows(values_only=True), start=1):
            if i <= 4 or r is None:
                continue
            ngay = r[0]
            if ngay is None or not hasattr(ngay, 'year'):
                continue  # dòng tổng, dòng chú thích
            debit = float(r[7] or 0)
            credit = float(r[8] or 0)
            if debit > 0 and credit > 0:
                rep.fail(4, f'{sname} dòng {i}: có cả Nợ và Có cùng lúc')
            rows.append({
                'sheet': sname, 'row': i,
                'posting_date': ngay.date() if isinstance(ngay, datetime) else ngay,
                'voucher_date': r[1].date() if isinstance(r[1], datetime) else r[1],
                'voucher_no': str(r[2] or '').strip(),
                'invoice_no': str(r[3] or '').strip() or None,
                'description': str(r[4] or '').strip() or None,
                'account': str(r[5] or '').strip(),
                'contra': str(r[6] or '').strip() or None,
                'debit': debit, 'credit': credit,
                'partner': str(r[9] or '').strip() or None,
                'partner_name': str(r[10] or '').strip() or None,
                'deductible': 'không hợp lý' not in str(r[11] or '').lower(),
            })
    wb.close()
    return rows, sheets


def read_trial_balance(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    sh = wb[wb.sheetnames[0]]
    tb = {}
    for i, r in enumerate(sh.iter_rows(values_only=True), start=1):
        if i < 10 or r is None or r[0] is None:
            continue
        code = str(r[0]).strip()
        if not code or not code[0].isdigit():
            continue
        def num(x):
            try: return float(x or 0)
            except (TypeError, ValueError): return 0.0
        tb[code] = {'debit': num(r[4]), 'credit': num(r[5])}
    wb.close()
    return tb


# ── chương trình chính ──────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('src', help='thư mục chứa bộ Excel')
    ap.add_argument('--emit-sql', metavar='FILE', help='sinh file SQL thay vì chỉ chạy thử')
    ap.add_argument('--report', default='finance-import-report.txt')
    ap.add_argument('--acknowledge', metavar='TK', nargs='*', default=[],
                    help='Mã tài khoản mà kế toán đã xác nhận chênh lệch là hợp lệ. '
                         'Phải liệt kê tường minh từng mã, không có tùy chọn bỏ qua tất cả.')
    args = ap.parse_args()

    rep = Report()
    S = lambda n: os.path.join(args.src, n)

    rep.w('=' * 78)
    rep.w('FINANCE VAULT · NHẬP BỘ SỔ KẾ TOÁN')
    rep.w('=' * 78)
    rep.w(f'Nguồn      : {args.src}')
    rep.w(f'Chế độ     : {"SINH SQL" if args.emit_sql else "CHẠY THỬ, không ghi gì"}')
    rep.w('')

    # ── tầng 1: cấu trúc ────────────────────────────────────────────────────
    rep.w('TẦNG 1 · CẤU TRÚC')
    need = {
        'accounts':  'Danh_sach_he_thong_tai_khoan.xlsx',
        'costs':     'Danh_sach_khoan_muc_chi_phi.xlsx',
        'suppliers': 'Danh_sach_nha_cung_cap.xlsx',
        'journal':   'So_nhat_ky_chung.xlsx',
        'tb':        'Bang_can_doi_tai_khoan.xlsx',
    }
    missing = [f for f in need.values() if not os.path.exists(S(f))]
    if missing:
        rep.w('   THIẾU FILE: ' + ', '.join(missing))
        rep.fail(1, 'thiếu file bắt buộc')
        rep.dump(args.report)
        sys.exit(1)
    for k, f in need.items():
        rep.w(f'   {f:<44} {os.path.getsize(S(f))//1024:>6} KB')
    rep.w('')

    # ── tầng 2: đọc và ép kiểu ──────────────────────────────────────────────
    rep.w('TẦNG 2 · ĐỌC VÀ ÉP KIỂU')
    accounts = read_accounts(S(need['accounts']), rep)
    costs    = read_cost_items(S(need['costs']), rep)
    suppliers= read_suppliers(S(need['suppliers']), rep)
    journal, sheets = read_journal(S(need['journal']), rep)
    tb       = read_trial_balance(S(need['tb']))
    rep.w(f'   tài khoản        : {len(accounts):>8,}')
    rep.w(f'   khoản mục chi phí: {len(costs):>8,}   (chi nhánh: '
          + ', '.join(sorted({c["branch"] for c in costs if c["branch"]})) + ')')
    rep.w(f'   nhà cung cấp     : {len(suppliers):>8,}')
    rep.w(f'   dòng nhật ký     : {len(journal):>8,}   ({len(sheets)} sheet)')
    rep.w('')

    # ── tầng 3: nghiệp vụ ───────────────────────────────────────────────────
    rep.w('TẦNG 3 · NGHIỆP VỤ')
    acc_codes = {a['code'] for a in accounts}
    unknown_acc = defaultdict(int)
    for r in journal:
        if r['account'] not in acc_codes:
            unknown_acc[r['account']] += 1
    if unknown_acc:
        rep.w(f'   tài khoản lạ: {len(unknown_acc)} mã')
        for c, n in sorted(unknown_acc.items(), key=lambda x: -x[1])[:8]:
            rep.w(f'      {c:<10} {n:,} dòng')
        rep.fail(3, f'{len(unknown_acc)} tài khoản không có trong danh mục')
    else:
        rep.w('   mọi tài khoản trong nhật ký đều có trong danh mục')

    # đối tượng: gộp từ nhà cung cấp + phát sinh trong nhật ký
    partners = {s['code']: s for s in suppliers}
    for r in journal:
        c = r['partner']
        if c and c not in partners:
            kind, branch = classify_partner(c)
            partners[c] = {'code': c, 'name': r['partner_name'] or c, 'kind': kind,
                           'branch': branch, 'address': None, 'tax_code': None, 'phone': None}
    kinds = defaultdict(int)
    for p in partners.values():
        kinds[p['kind']] += 1
    rep.w(f'   đối tượng công nợ: {len(partners):,}  ('
          + ', '.join(f'{k} {v:,}' for k, v in sorted(kinds.items())) + ')')
    rep.w('')

    # ── tầng 4: bất biến kế toán ────────────────────────────────────────────
    rep.w('TẦNG 4 · BẤT BIẾN KẾ TOÁN')
    tot_d = sum(r['debit'] for r in journal)
    tot_c = sum(r['credit'] for r in journal)
    rep.w(f'   tổng phát sinh Nợ : {money(tot_d):>22}')
    rep.w(f'   tổng phát sinh Có : {money(tot_c):>22}')
    rep.w(f'   lệch              : {money(tot_d - tot_c):>22}')
    if abs(tot_d - tot_c) >= 1:
        rep.fail(4, f'Nợ khác Có: lệch {money(tot_d-tot_c)}')
        rep.w('   KHÔNG CÂN BẰNG · dừng, không nhập')
    else:
        rep.w('   CÂN BẰNG')

    # nhóm chứng từ
    by_doc = defaultdict(list)
    for r in journal:
        by_doc[(r['posting_date'], r['voucher_no'])].append(r)
    unbal = []
    for k, g in by_doc.items():
        d = sum(x['debit'] for x in g)
        c = sum(x['credit'] for x in g)
        if abs(d - c) >= 1:
            unbal.append((k, d - c, len(g)))
    rep.w(f'   chứng từ          : {len(by_doc):,}')
    rep.w(f'   chứng từ tự cân   : {len(by_doc)-len(unbal):,}')
    rep.w(f'   cân theo cặp      : {len(unbal):,}')

    # ghép cặp các chứng từ lệch: hai chứng từ cùng ngày, lệch triệt tiêu nhau
    groups = {}
    used = set()
    for i, (k1, d1, _) in enumerate(unbal):
        if k1 in used:
            continue
        for k2, d2, _ in unbal[i+1:]:
            if k2 in used or k2[0] != k1[0]:
                continue
            if abs(d1 + d2) < 1:
                gid = f'{k1[0].isoformat()}:{k1[1]}+{k2[1]}'
                groups[k1] = gid
                groups[k2] = gid
                used.add(k1); used.add(k2)
                break
    paired = len(groups)
    rep.w(f'   ghép được thành cặp: {paired:,} / {len(unbal):,}')
    if paired < len(unbal):
        rep.fail(4, f'{len(unbal)-paired} chứng từ lệch không ghép được cặp')
        for k, d, n in unbal:
            if k not in groups:
                rep.w(f'      KHÔNG GHÉP ĐƯỢC: {k[0]} {k[1]} lệch {money(d)}')
    rep.w('')

    # ── tầng 5: đối chiếu chéo ──────────────────────────────────────────────
    rep.w('TẦNG 5 · ĐỐI CHIẾU CHÉO VỚI BẢNG CÂN ĐỐI TÀI KHOẢN')
    ja = defaultdict(lambda: [0.0, 0.0])
    for r in journal:
        ja[r['account']][0] += r['debit']
        ja[r['account']][1] += r['credit']
    both = sorted(set(ja) & set(tb))
    ok, diff = [], []
    for c in both:
        if abs(ja[c][0] - tb[c]['debit']) < 1 and abs(ja[c][1] - tb[c]['credit']) < 1:
            ok.append(c)
        else:
            diff.append((c, ja[c][0], tb[c]['debit'], ja[c][1], tb[c]['credit']))
    rep.w(f'   tài khoản đối chiếu được : {len(both)}')
    rep.w(f'   KHỚP                     : {len(ok)}')
    rep.w(f'   LỆCH                     : {len(diff)}')
    for c, jd, bd, jc, bc in diff:
        rep.w(f'      TK {c:<7} Nợ {money(jd):>18} / {money(bd):>18}'
              f'   Có {money(jc):>18} / {money(bc):>18}')
    ack = set(args.acknowledge)
    unack = [d for d in diff if d[0] not in ack]
    if ack:
        rep.w('')
        rep.w('   Kế toán đã xác nhận chênh lệch hợp lệ ở: ' + ', '.join(sorted(ack)))
        for c in sorted(ack):
            if c not in {d[0] for d in diff}:
                rep.w(f'      lưu ý: TK {c} được xác nhận nhưng thực tế không lệch')
    if unack:
        rep.fail(5, f'{len(unack)} tài khoản lệch chưa được xác nhận: '
                    + ', '.join(d[0] for d in unack))
    rep.w('')

    # ── kết luận ────────────────────────────────────────────────────────────
    rep.w('=' * 78)
    if rep.errors:
        rep.w(f'KẾT LUẬN · CHẶN LÔ · {len(rep.errors)} vấn đề')
        for tier, msg in rep.errors:
            rep.w(f'   tầng {tier}: {msg}')
        rep.w('')
        rep.w('Theo quy trình, lô này không được ghi vào sổ cái cho tới khi')
        rep.w('kế toán giải thích được các chênh lệch trên.')
    else:
        rep.w('KẾT LUẬN · ĐẠT CẢ NĂM TẦNG · sẵn sàng để kế toán duyệt')
    rep.w('=' * 78)

    if args.emit_sql:
        if rep.errors:
            rep.w('')
            rep.w('Không sinh SQL vì lô chưa đạt kiểm tra.')
        else:
            import json
            recon_json = json.dumps({
                'total_debit': tot_d, 'total_credit': tot_c,
                'vouchers': len(by_doc), 'lines': len(journal),
                'accounts_matched': len(ok), 'accounts_differing': len(diff),
                'acknowledged': sorted(ack),
                'differences': [{'account': c, 'journal_debit': jd, 'tb_debit': bd,
                                 'journal_credit': jc, 'tb_credit': bc}
                                for c, jd, bd, jc, bc in diff],
            }, ensure_ascii=False)
            emit(args.emit_sql, accounts, costs, partners, journal, by_doc, groups,
                 sha256(S(need['journal'])), need['journal'], recon_json)
            rep.w('')
            rep.w(f'Đã sinh SQL: {args.emit_sql}')

    rep.dump(args.report)
    print(f'\nBáo cáo đầy đủ: {args.report}')
    sys.exit(1 if rep.errors else 0)


def emit(path, accounts, costs, partners, journal, by_doc, groups, digest, fname, recon_json='{}'):
    """Sinh SQL nạp. Bọc trong một giao dịch, hỏng là hoàn tác sạch."""
    per = sorted({r['posting_date'].strftime('%Y-%m') for r in journal})
    with open(path, 'w', encoding='utf-8') as f:
        W = lambda s='': f.write(s + '\n')
        W('-- Sinh tự động bởi scripts/finance-import.py. Không sửa tay.')
        W(f'-- Nguồn: {fname}  sha256={digest}')
        W('begin;')
        W("insert into finance.import_batches(id, source_file, source_sha256, row_count, status, recon)")
        W(f"values (gen_random_uuid(), {sql_str(fname)}, {sql_str(digest)}, {len(journal)}, 'validated',")
        W(f"        {sql_str(recon_json)}::jsonb);")
        W("create temp table _b as select id from finance.import_batches "
          f"where source_sha256 = {sql_str(digest)} order by created_at desc limit 1;")
        W()
        for p in per:
            y, m = p.split('-')
            W(f"insert into finance.periods(code, start_date, end_date) values "
              f"('{p}', '{p}-01', (date '{p}-01' + interval '1 month - 1 day')::date) "
              f"on conflict (code) do nothing;")
        W()
        for a in accounts:
            W("insert into finance.accounts(code,name,name_en,nature,depth,is_active,note) values ("
              f"{sql_str(a['code'])},{sql_str(a['name'])},{sql_str(a['name_en'])},"
              f"{sql_str(a['nature'])},{a['depth']},{str(a['is_active']).lower()},{sql_str(a['note'])}) "
              "on conflict (code) do update set name=excluded.name, updated_at=now();")
        W()
        for c in costs:
            W("insert into finance.cost_items(code,name,branch_code,is_active) values ("
              f"{sql_str(c['code'])},{sql_str(c['name'])},{sql_str(c['branch'])},"
              f"{str(c['is_active']).lower()}) on conflict (code) do nothing;")
        W()
        for p in partners.values():
            W("insert into finance.partners(code,name,kind,branch_hint,tax_code,address,phone) values ("
              f"{sql_str(p['code'])},{sql_str(p['name'])},{sql_str(p['kind'])},"
              f"{sql_str(p.get('branch'))},{sql_str(p.get('tax_code'))},"
              f"{sql_str(p.get('address'))},{sql_str(p.get('phone'))}) "
              "on conflict (code) do nothing;")
        W()
        for (pdate, vno), lines in by_doc.items():
            head = lines[0]
            vtype = (VOUCHER_TYPE.match(vno).group(1).upper() if VOUCHER_TYPE.match(vno) else None)
            gid = groups.get((pdate, vno))
            W("insert into finance.vouchers(voucher_no,posting_date,voucher_date,voucher_type,"
              "invoice_no,description,period_code,balance_group,batch_id,source_ref) values ("
              f"{sql_str(vno)},{sql_date(pdate)},{sql_date(head['voucher_date'])},{sql_str(vtype)},"
              f"{sql_str(head['invoice_no'])},{sql_str(head['description'])},"
              f"'{pdate.strftime('%Y-%m')}',{sql_str(gid)},(select id from _b),"
              f"""'{{"file":"{fname}"}}'::jsonb) on conflict (voucher_no, posting_date) do nothing;""")
            for n, l in enumerate(lines, start=1):
                W("insert into finance.journal_lines(voucher_id,line_no,account_code,"
                  "contra_account_code,debit,credit,partner_code,description,is_deductible,"
                  "source_sheet,source_row) select v.id,"
                  f"{n},{sql_str(l['account'])},{sql_str(l['contra'])},{l['debit']:.2f},{l['credit']:.2f},"
                  f"{sql_str(l['partner'])},{sql_str(l['description'])},{str(l['deductible']).lower()},"
                  f"{sql_str(l['sheet'])},{l['row']} from finance.vouchers v "
                  f"where v.voucher_no={sql_str(vno)} and v.posting_date={sql_date(pdate)};")
        W()
        W("update finance.import_batches set status='posted', posted_at=now() "
          "where id = (select id from _b);")
        W('commit;')


if __name__ == '__main__':
    main()
